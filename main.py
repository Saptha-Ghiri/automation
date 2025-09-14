import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from copy import copy
import xlwings as xw
import xlsxwriter
import json
import os
import tempfile
import time
import threading
from datetime import datetime, timedelta
import re
from extract_queue_data import extract_resource_status_counts, create_sample_data
from ppt_automation import generate_weekly_report
status_str = None

def cleanup_temp_files():
    """Securely delete temporary files"""
    if 'temp_files' in st.session_state:
        for temp_file in st.session_state.temp_files:
            try:
                if os.path.exists(temp_file):
                    os.remove(temp_file)
                    st.info(f"🧹 Cleaned up temporary file: {temp_file}")
            except Exception as e:
                st.warning(f"Could not remove temp file {temp_file}: {e}")
        st.session_state.temp_files = []

# Security warning function removed - users assumed to accept risk
def extract_date_period_from_excel(file_path):
    """Extract date period from Excel file cell B7"""
    try:
        wb = load_workbook(file_path)
        # Try to get the 'Cloud Services Report' sheet first, if not found use first sheet
        if 'Cloud Services Report' in wb.sheetnames:
            ws = wb['Cloud Services Report']
        else:
            ws = wb.active
        
        # Get cell B7 content
        b7_value = ws['B7'].value
        
        if b7_value:
            # Convert to string if not already
            date_text = str(b7_value)
            st.info(f"Found date field in B7: {date_text}")
            
            # Extract date range using regex - looking for pattern like (9/1/2025 to 9/7/2025)
            # Try multiple patterns to handle different formats
            patterns = [
                r'\((\d{1,2}/\d{1,2}/\d{4})\s+to\s+(\d{1,2}/\d{1,2}/\d{4})\)',  # (9/1/2025 to 9/7/2025)
                r'(\d{1,2}/\d{1,2}/\d{4})\s+to\s+(\d{1,2}/\d{1,2}/\d{4})',      # 9/1/2025 to 9/7/2025
                r'Custom\s*\((\d{1,2}/\d{1,2}/\d{4})\s+to\s+(\d{1,2}/\d{1,2}/\d{4})\)',  # Custom (9/1/2025 to 9/7/2025)
                r'equals\s*Custom\s*\((\d{1,2}/\d{1,2}/\d{4})\s+to\s+(\d{1,2}/\d{1,2}/\d{4})\)' # equals Custom (9/1/2025 to 9/7/2025)
            ]
            
            match = None
            for pattern in patterns:
                match = re.search(pattern, date_text, re.IGNORECASE)
                if match:
                    st.success(f"✅ Found date pattern: {pattern}")
                    break
            
            if match:
                start_date_str = match.group(1)
                end_date_str = match.group(2)
                
                # Parse dates
                start_date = datetime.strptime(start_date_str, '%m/%d/%Y')
                end_date = datetime.strptime(end_date_str, '%m/%d/%Y')
                
                # Calculate report date (end date + 1 day)
                report_date = end_date + timedelta(days=1)
                
                # Format strings
                period_str = f"{start_date.strftime('%m/%d/%Y')} to {end_date.strftime('%m/%d/%Y')}"
                report_date_str = report_date.strftime('%d %B %Y')  # e.g., "8 September 2025"
                new_date_str = end_date.strftime('%m/%d/%Y')  # e.g., "09/07/2025"
                
                return {
                    'period': period_str,
                    'report_date': report_date_str,
                    'new_date': new_date_str,
                    'start_date': start_date,
                    'end_date': end_date,
                    'report_date_obj': report_date
                }
            else:
                st.warning(f"Could not parse date pattern from: {date_text}")
        else:
            st.warning("Cell B7 is empty")
            
    except Exception as e:
        st.error(f"Error extracting date period: {e}")
    
    # Return default values if extraction fails
    return {
        'period': '09/01/2025 to 09/07/2025',
        'report_date': '8 September 2025',
        'new_date': '09/07/2025',
        'start_date': datetime(2025, 9, 1),
        'end_date': datetime(2025, 9, 7),
        'report_date_obj': datetime(2025, 9, 8)
    }

def safe_str(value):
    """Convert value to string safely, handling Unicode characters"""
    if value is None:
        return ""
    try:
        return str(value)
    except UnicodeEncodeError:
        return str(value).encode('ascii', errors='ignore').decode('ascii')

def check_and_cleanup_empty_section():
    """Check if current section is empty and remove subtotal if needed"""
    ws = st.session_state.ws
    sections = st.session_state.sections
    current_section = st.session_state.current_section
    
    if current_section >= len(sections) - 1:
        return False  # No more sections
    
    start_row = sections[current_section] + 1  # First row after header/previous subtotal
    end_row = sections[current_section + 1]    # Subtotal row of current section
    
    # Check if section has any valid tickets (non-subtotal, non-empty rows)
    has_tickets = False
    for row in range(start_row, end_row):
        # Skip if this is the subtotal row itself
        val_status = ws.cell(row=row, column=2).value
        if val_status and str(val_status).strip() in ["Subtotal", "Total"]:
            continue
            
        # Skip count rows  
        val_col3 = ws.cell(row=row, column=3).value
        if val_col3 and str(val_col3).strip() == "Count":
            continue
            
        # Check if row has ticket data
        val_case_number = ws.cell(row=row, column=4).value
        val_subject = ws.cell(row=row, column=7).value
        val_responsible = ws.cell(row=row, column=5).value
        
        if val_case_number or val_subject or val_responsible:
            has_tickets = True
            break
    
    # If no tickets found, delete the subtotal row
    if not has_tickets:
        subtotal_row = end_row  # This is the subtotal row
        ws.delete_rows(subtotal_row)
        st.session_state.deleted_rows += 1
        
        # Update sections list to remove this empty section
        st.session_state.sections.pop(current_section + 1)
        
        # Don't increment current_section since we removed a section
        return True  # Section was cleaned up
    
    return False  # Section still has tickets

# Set page config
st.set_page_config(
    page_title="CSM Report Processor",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for report generation section only
st.markdown("""
<style>
    /* Statistics cards for report generation */
    .stat-card {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        padding: 1.5rem;
        border-radius: 10px;
        text-align: center;
        margin: 0.5rem 0;
    }
    
    /* Card styling for report generation section */
    .report-card {
        background: white;
        padding: 1.5rem;
        border-radius: 10px;
        box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        border-left: 4px solid #667eea;
        margin-bottom: 1rem;
    }
</style>
""", unsafe_allow_html=True)

# Initialize session state - merged from main.py
if 'file_processed' not in st.session_state:
    st.session_state.file_processed = False
if 'r' not in st.session_state:
    st.session_state.r = [13]
if 'current_row' not in st.session_state:
    st.session_state.current_row = 13
if 'total' not in st.session_state:
    st.session_state.total = 0
if 'processing_complete' not in st.session_state:
    st.session_state.processing_complete = False
if 'stats' not in st.session_state:
    st.session_state.stats = {
        'priority': {"Priority 1": 0, "Priority 2": 0, "Priority 3": 0, "Priority 4": 0},
        'account_count': {"Automic": 0, "Beigene": 0, "BMS": 0, "Collegium": 0, 
                         "Azure Imdaas": 0, "Aws Imdaas": 0, "MDM": 0, "Usbu-Pede": 0},
        'dict_status': {"New": 0, "Inprogress": 0, "Awaiting": 0, 
                       "Internal Solution Provided": 0, "Resolved with Customer": 0, "Closed": 0},
        'ticket_completed': {"Abhijeet Nashikkar": 0, "Aditya Anand": 0, 
                            "Nishanth Senthilkumar": 0, "Sakthivel s Venkatachalam": 0},
        'sla': {"SLA Met": 100, "SLA Lost": 0}
    }
if 'wb' not in st.session_state:
    st.session_state.wb = None
if 'ws' not in st.session_state:
    st.session_state.ws = None
if 'file_path' not in st.session_state:
    st.session_state.file_path = None
if 'temp_daas_file' not in st.session_state:
    st.session_state.temp_daas_file = None
if 'temp_daas_processed' not in st.session_state:
    st.session_state.temp_daas_processed = False
if 'temp_daas_data' not in st.session_state:
    st.session_state.temp_daas_data = None
if 'ppt_generated' not in st.session_state:
    st.session_state.ppt_generated = False
if 'combined_json_data' not in st.session_state:
    st.session_state.combined_json_data = None
if 'combined_json_path' not in st.session_state:
    st.session_state.combined_json_path = None
if 'date_info' not in st.session_state:
    st.session_state.date_info = None

def add_horizontal_chart(file_path, sheet_name, start_row, start_col, chart_title="Chart", chart_type="bar_clustered", chart_index=0):
    """Create charts using xlwings"""
        
    app = xw.App(visible=False)
    try:
        wb = app.books.open(file_path)
        sheet = wb.sheets[sheet_name]

        # Find last row of data
        last_row = sheet.range((start_row, start_col)).end('down').row
        last_col = start_col + 1

        # Position chart
        anchor_cell = sheet.range((start_row, last_col))
        chart_left = anchor_cell.left + anchor_cell.width + 20
        chart_top = anchor_cell.top + (chart_index * 40)

        data_range = sheet.range(sheet.cells(start_row, start_col), sheet.cells(last_row, last_col))

        chart = sheet.charts.add(left=chart_left, top=chart_top)
        chart.name = f"{chart_title}_{chart_index}"
        chart.chart_type = chart_type
        chart.width = 300
        chart.height = 200
        chart.set_source_data(data_range)

        chart_api = chart.api[1]
        chart_api.HasTitle = True
        chart_api.ChartTitle.Text = chart_title
        chart_api.ChartTitle.Font.Size = 10
        chart_api.HasDataTable = True
        chart_api.DataTable.ShowLegendKey = True
        chart_api.DataTable.Font.Size = 8
        chart_api.HasLegend = True
        chart_api.Legend.Position = -4107
        chart_api.Legend.Font.Size = 8
        chart_api.ChartGroups(1).GapWidth = 200
        chart_api.ChartGroups(1).Overlap = 0
        chart_api.Axes(1).TickLabels.Font.Size = 8
        chart_api.Axes(2).TickLabels.Font.Size = 8

        wb.save()
        wb.close()
        return True
    except Exception as e:
        st.error(f"Error creating chart: {e}")
        return False
    finally:
        app.quit()

def add_pie_chart(file_path, sheet_name, start_row, start_col, chart_title, chart_index=0):
    """Create pie chart using xlwings"""
        
    app = xw.App(visible=False)
    try:
        wb = app.books.open(file_path)
        sheet = wb.sheets[sheet_name]

        # Find last row of data
        last_row = sheet.range((start_row, start_col)).end('down').row
        last_col = start_col + 1

        # Position chart
        anchor_cell = sheet.range((start_row, last_col))
        chart_left = anchor_cell.left + anchor_cell.width + 20
        chart_top = anchor_cell.top + (chart_index * 40)

        data_range = sheet.range((start_row, start_col), (last_row, last_col))

        chart = sheet.charts.add(left=chart_left, top=chart_top)
        chart.name = f"{chart_title}_{chart_index}"
        chart.chart_type = 'pie'
        chart.width = 250
        chart.height = 200
        chart.set_source_data(data_range)

        # Enhanced pie chart formatting
        try:
            chart_api = chart.api[1]
            chart_api.HasTitle = True
            chart_api.ChartTitle.Text = chart_title
            
            series = chart_api.SeriesCollection(1)
            series.HasDataLabels = True
            data_labels = series.DataLabels
            data_labels.ShowPercentage = True
            data_labels.ShowValue = True
            data_labels.ShowCategoryName = True
            data_labels.Position = -4142
            
            chart_api.HasLegend = True
            legend = chart_api.Legend
            legend.Position = -4107
            
        except Exception as format_error:
            st.warning(f"Could not apply advanced formatting: {format_error}")

        wb.save()
        wb.close()
        return True
    except Exception as e:
        st.error(f"Error creating pie chart: {e}")
        return False
    finally:
        app.quit()

def add_horizontal_chart_xlwings(
    sheet,
    start_row,
    start_col,
    chart_title="Chart",
    chart_type="bar_clustered",
    width=300,
    height=200,
    font_size=8,
    chart_index=0
):
    """Add horizontal bar chart to xlwings sheet - from main.py"""
    try:
        last_row = sheet.range((start_row, start_col)).end('down').row
        last_col = sheet.range((start_row, start_col)).end('right').column

        anchor_cell = sheet.range((start_row, last_col))
        chart_left = anchor_cell.left + anchor_cell.width + 20
        chart_top = anchor_cell.top + (chart_index * 50)

        data_range = sheet.range(sheet.cells(start_row, start_col), sheet.cells(last_row, last_col))

        chart = sheet.charts.add(left=chart_left, top=chart_top)
        chart.name = f"{chart_title}_{chart_index}"
        
        if chart_type == "pie":
            chart.chart_type = "pie"
        else:
            chart.chart_type = chart_type
            
        chart.width = width
        chart.height = height
        chart.set_source_data(data_range)

        chart_api = chart.api[1]
        chart_api.HasTitle = True
        chart_api.ChartTitle.Text = chart_title
        chart_api.ChartTitle.Font.Size = font_size + 2
        chart_api.HasDataTable = True
        chart_api.DataTable.ShowLegendKey = True
        chart_api.DataTable.Font.Size = font_size
        chart_api.HasLegend = True
        chart_api.Legend.Position = -4107
        chart_api.Legend.Font.Size = font_size
        chart_api.ChartGroups(1).GapWidth = 200
        chart_api.ChartGroups(1).Overlap = 0
        chart_api.Axes(1).TickLabels.Font.Size = font_size
        chart_api.Axes(2).TickLabels.Font.Size = font_size

        st.success(f"Chart '{chart_title}' added successfully!")
    except Exception as e:
        st.error(f"Error creating chart: {e}")

def add_pie_chart_xlwings(sheet, start_row, start_col, chart_title, chart_index=0):
    """Add pie chart to xlwings sheet - from main.py"""
    try:
        last_row = sheet.range((start_row, start_col)).end('down').row
        last_col = start_col + 1

        anchor_cell = sheet.range((start_row, last_col))
        chart_left = anchor_cell.left + anchor_cell.width + 20
        chart_top = anchor_cell.top + (chart_index * 50)

        data_range = sheet.range((start_row, start_col), (last_row, last_col))

        chart = sheet.charts.add(left=chart_left, top=chart_top)
        chart.name = f"{chart_title}_{chart_index}"
        chart.chart_type = 'pie'
        chart.width = 250
        chart.height = 200
        chart.set_source_data(data_range)

        try:
            chart_api = chart.api[1]
            chart_api.HasTitle = True
            chart_api.ChartTitle.Text = chart_title
            
            series = chart_api.SeriesCollection(1)
            series.HasDataLabels = True
            data_labels = series.DataLabels
            data_labels.ShowPercentage = True
            data_labels.ShowValue = True
            data_labels.ShowCategoryName = True
            data_labels.Position = -4142
            
            chart_api.HasLegend = True
            legend = chart_api.Legend
            legend.Position = -4107
            st.success("Pie chart enhanced with data labels and legend")
        except Exception as format_error:
            st.warning(f"Warning: Could not apply advanced formatting: {format_error}")

        st.success(f"Pie chart '{chart_title}' added successfully!")
    except Exception as e:
        st.error(f"Error creating pie chart: {e}")

def process_temp_daas_file(temp_daas_file):
    """Process temp_daas_queue file in background"""
    if temp_daas_file is not None:
        # Save temp file with unique name
        import uuid
        unique_id = str(uuid.uuid4())[:8]
        temp_daas_path = f"temp_daas_queue_{unique_id}.xlsx"
        
        try:
            with open(temp_daas_path, "wb") as f:
                f.write(temp_daas_file.getvalue())
            
            # Store temp file path for cleanup later
            if 'temp_files' not in st.session_state:
                st.session_state.temp_files = []
            st.session_state.temp_files.append(temp_daas_path)
            
            # Extract data using existing function
            resource_counts, status_counts, date_wise_data = extract_resource_status_counts(temp_daas_path)
        except Exception as e:
            st.error(f"Error processing DaaS queue file: {e}")
            return None
        
        # If extraction failed, use sample data
        if resource_counts is None:
            resource_counts, status_counts, date_wise_data = create_sample_data()
        
        return {
            'resource_counts': resource_counts,
            'status_counts': status_counts,
            'date_wise_data': date_wise_data
        }
    return None

def process_uploaded_file(uploaded_file):
    """Process uploaded file using simplified logic from main.py"""
    # Save uploaded file as working file with secure temp file
    import tempfile
    import uuid
    
    # Create unique temp file name to avoid conflicts
    unique_id = str(uuid.uuid4())[:8]
    temp_file_path = f"working_file_{unique_id}.xlsx"
    
    try:
        with open(temp_file_path, "wb") as f:
            f.write(uploaded_file.getvalue())
        
        # Store temp file path for cleanup later
        if 'temp_files' not in st.session_state:
            st.session_state.temp_files = []
        st.session_state.temp_files.append(temp_file_path)
        
        st.session_state.file_path = temp_file_path
    except Exception as e:
        st.error(f"Error saving uploaded file: {e}")
        return
    st.session_state.wb = load_workbook(temp_file_path)
    st.session_state.ws = st.session_state.wb['Cloud Services Report']
    
    # Extract date period from B7 cell
    date_info = extract_date_period_from_excel(temp_file_path)
    st.session_state.date_info = date_info
    
    st.success(f"📅 Extracted Period: {date_info['period']}")
    st.success(f"📊 Report Date: {date_info['report_date']}")

    ws = st.session_state.ws
    
    # Insert columns and formatting - same as main.py
    insert_after = 7
    ws.insert_cols(insert_after + 1, amount=2)
    for row in range(1, ws.max_row + 1):
        for src_col, dest_col in zip([5, 6], [insert_after + 1, insert_after + 2]):
            src_cell = ws.cell(row=row, column=src_col)
            dest_cell = ws.cell(row=row, column=dest_col)
            
            dest_cell.value = src_cell.value
            
            if src_cell.has_style:
                dest_cell.font = copy(src_cell.font)
                dest_cell.border = copy(src_cell.border)
                dest_cell.fill = copy(src_cell.fill)
                dest_cell.number_format = copy(src_cell.number_format)
                dest_cell.protection = copy(src_cell.protection)
                dest_cell.alignment = copy(src_cell.alignment)

    for src_col, dest_col in zip([5, 6], [insert_after + 1, insert_after + 2]):
        letter_src = ws.cell(row=1, column=src_col).column_letter
        letter_dest = ws.cell(row=1, column=dest_col).column_letter
        ws.column_dimensions[letter_dest].width = ws.column_dimensions[letter_src].width

    ws.cell(row=12, column=8, value="Actions")
    ws.cell(row=12, column=9, value="Account")
    
    st.session_state.wb.save(st.session_state.file_path)
    st.session_state.file_processed = True

def get_current_ticket_for_processing():
    """Get current ticket details using simplified logic from main.py"""
    ws = st.session_state.ws
    search_word = "Subtotal"
    col_B = 2
    
    row = st.session_state.current_row
    max_rows = ws.max_row
    
    if row <= max_rows:
        val_b = ws.cell(row=row, column=col_B).value
        
        if val_b and str(val_b).strip() == search_word:
            return "subtotal_found"
        elif val_b and str(val_b).strip() == "Total":
            return "total_found" 
        else:
            # Valid ticket found - get ticket data
            status = ws.cell(row=row, column=2).value
            user = ws.cell(row=row, column=5).value  
            priority = ws.cell(row=row, column=12).value
            subject = ws.cell(row=row, column=7).value
            return {
                'row': row,
                'status': safe_str(status),
                'user': safe_str(user),
                'subject': safe_str(subject),
                
            }
    else:
        return None

def check_and_cleanup_empty_section_after_delete(deleted_row):
    """Check if deleting this row leaves a section empty, and remove subtotal if needed"""
    ws = st.session_state.ws
    
    # Find the current section this row belongs to
    # Look backwards from deleted_row to find the section header
    current_section_start = None
    for check_row in range(deleted_row - 1, 0, -1):
        cell_val = ws.cell(row=check_row, column=2).value
        if cell_val and str(cell_val).strip() not in ["Subtotal", "Total"]:
            # This might be a section header - check if it's different from the status we're in
            break
        elif cell_val and str(cell_val).strip() == "Subtotal":
            current_section_start = check_row + 1
            break
    
    if current_section_start is None:
        current_section_start = 13  # Default start
    
    # Find the next subtotal (end of current section)  
    section_end = None
    for check_row in range(deleted_row, ws.max_row + 1):
        cell_val = ws.cell(row=check_row, column=2).value
        if cell_val and str(cell_val).strip() == "Subtotal":
            section_end = check_row
            break
        elif cell_val and str(cell_val).strip() == "Total":
            section_end = check_row
            break
    
    if section_end is None:
        return False  # No section end found
    
    # Check if there are any remaining tickets between section start and subtotal
    has_remaining_tickets = False
    for check_row in range(current_section_start, section_end):
        if check_row == deleted_row:
            continue  # Skip the deleted row
            
        # Check for actual ticket data
        case_num = ws.cell(row=check_row, column=4).value
        subject = ws.cell(row=check_row, column=7).value
        responsible = ws.cell(row=check_row, column=5).value
        status = ws.cell(row=check_row, column=2).value
        
        # Skip empty rows and count/subtotal rows
        if status and str(status).strip() in ["Count", "Subtotal", "Total"]:
            continue
            
        if case_num or subject or responsible:
            has_remaining_tickets = True
            break
    
    # If no remaining tickets, delete the subtotal row
    if not has_remaining_tickets:
        ws.delete_rows(section_end)
        st.info(f"🧹 Section became empty - subtotal row {section_end} removed automatically!")
        return True
    
    return False

def update_section_subtotal_count(section_start_row, section_end_row):
    """Update the count for a specific section's subtotal"""
    ws = st.session_state.ws
    
    # Count actual tickets in this section
    ticket_count = 0
    for check_row in range(section_start_row, section_end_row):
        # Check for actual ticket data (not count/subtotal rows)
        case_num = ws.cell(row=check_row, column=4).value
        subject = ws.cell(row=check_row, column=7).value
        responsible = ws.cell(row=check_row, column=5).value
        status = ws.cell(row=check_row, column=2).value
        
        # Skip empty rows and count/subtotal rows
        if status and str(status).strip() in ["Count", "Subtotal", "Total"]:
            continue
            
        if case_num or subject or responsible:
            ticket_count += 1
    
    # Update the subtotal row - find the "Count" cell in column 3
    subtotal_row = section_end_row
    for check_row in range(section_end_row - 2, section_end_row + 1):
        if check_row > ws.max_row:
            continue
        cell_val = ws.cell(row=check_row, column=3).value
        if cell_val and str(cell_val).strip() == "Count":
            ws.cell(row=check_row, column=4, value=ticket_count)
            st.info(f"📊 Section subtotal updated to {ticket_count} tickets")
            break
    
    return ticket_count

def update_all_subtotals_and_total():
    """Update all section subtotals and the final total"""
    ws = st.session_state.ws
    total_tickets = 0
    
    # Find all subtotal sections
    subtotal_rows = []
    for row in range(13, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=2).value
        if cell_val and str(cell_val).strip() == "Subtotal":
            subtotal_rows.append(row)
    
    # Find total row
    total_row = None
    for row in range(13, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=2).value
        if cell_val and str(cell_val).strip() == "Total":
            total_row = row
            break
    
    # Update each section
    section_start = 13  # First section starts at row 13
    for subtotal_row in subtotal_rows:
        section_count = update_section_subtotal_count(section_start, subtotal_row)
        total_tickets += section_count
        section_start = subtotal_row + 1  # Next section starts after this subtotal
    
    # Update the final total
    if total_row:
        # Find the count cell for the total (usually in column 4)
        ws.cell(row=total_row, column=4, value=total_tickets)
        st.success(f"🎯 Final total updated to {total_tickets} tickets")
    
    # Update session state total for chart generation
    st.session_state.total = total_tickets
    
    return total_tickets

def process_current_ticket(action, action_text="", selected_account=""):
    """Process ticket using simplified logic from main.py"""
    ws = st.session_state.ws
    row = st.session_state.current_row
    
    if action == "delete":
        # Check if we need to cleanup empty section BEFORE deleting the row
        section_cleaned = check_and_cleanup_empty_section_after_delete(row)
        
        # Delete the ticket row
        ws.delete_rows(row)
        st.success(f"Row {row} deleted.")
        
        # If section was cleaned up, we deleted an extra row, so don't advance current_row
        if section_cleaned:
            # Section subtotal was also deleted, so current_row should stay the same
            pass
        else:
            # Section still exists, update its subtotal and the grand total
            update_all_subtotals_and_total()
        
        st.session_state.wb.save(st.session_state.file_path)
        st.session_state.current_row = row  # Stay at same row (next row shifts up)
    else:
        # Update row data
        ws.cell(row=row, column=8, value=action_text)
        ws.cell(row=row, column=9, value=selected_account)
        
        # Update statistics
        user_name = ws.cell(row=row, column=5).value
        priority_val = ws.cell(row=row, column=12).value  
        status_val = ws.cell(row=row, column=2).value
        
        # Initialize status_str in session state if it doesn't exist
        if 'status_str' not in st.session_state:
            st.session_state.status_str = None
        
        if selected_account in st.session_state.stats['account_count']:
            st.session_state.stats['account_count'][selected_account] += 1
        if user_name and str(user_name) in st.session_state.stats['ticket_completed']:
            st.session_state.stats['ticket_completed'][str(user_name)] += 1
        if priority_val and str(priority_val) in st.session_state.stats['priority']:
            st.session_state.stats['priority'][str(priority_val)] += 1
        print(f"Status Value: {status_val}, {type(status_val)}")  # Debugging line
        print(f"Status String before assignment: {st.session_state.status_str}")  # Debugging line
        if status_val is None:
            if st.session_state.status_str is None:
                st.session_state.status_str = "consider as previous stats as current"
            # else keep the previous status_str value
        else:
            st.session_state.status_str = str(status_val)
        print(f"Status String after assignment: {st.session_state.status_str}")  # Debugging line
        if st.session_state.status_str in st.session_state.stats['dict_status']:
            st.session_state.stats['dict_status'][st.session_state.status_str] += 1
        print()
        st.session_state.total += 1
        st.success(f"Row {row} updated successfully.")
        
        # Update subtotals and total count after processing a ticket
        update_all_subtotals_and_total()
        
        st.session_state.wb.save(st.session_state.file_path)
        st.session_state.current_row += 1

def generate_charts_with_openpyxl():
    """Generate charts using openpyxl while preserving ALL original styles, fonts, colors"""
    try:
        stats = st.session_state.stats
        
        # Load the original file with openpyxl to preserve ALL formatting
        wb = load_workbook(st.session_state.file_path)
        ws = wb.active
        
        # Find the last row with data
        last_row = ws.max_row
        chart_start_row = last_row + 3
        col_offset = 5  # Column E (1-indexed = 5)
        
        # Add chart data with preserved styles
        from openpyxl.chart import BarChart, PieChart, Reference
        from openpyxl.chart.label import DataLabelList
        
        # 1. Ticket Status Chart Data
        ws.cell(row=chart_start_row, column=col_offset, value="TICKET STATUS")
        row_offset = chart_start_row + 1
        ws.cell(row=row_offset, column=col_offset, value="Status")
        ws.cell(row=row_offset, column=col_offset + 1, value="Count")
        
        for i, (status, count) in enumerate(stats['dict_status'].items()):
            ws.cell(row=row_offset + i + 1, column=col_offset, value=status)
            ws.cell(row=row_offset + i + 1, column=col_offset + 1, value=count)
        
        # Create horizontal bar chart for status with gap width
        chart1 = BarChart()
        chart1.type = "bar"
        chart1.style = 2  # Use a built-in style that includes blue colors
        chart1.title = "Ticket Status Count"
        chart1.y_axis.title = 'Status'
        chart1.x_axis.title = 'Count'
        
        data_end_row = row_offset + len(stats['dict_status'])
        data = Reference(ws, min_col=col_offset + 1, min_row=row_offset + 1, max_row=data_end_row, max_col=col_offset + 1)
        cats = Reference(ws, min_col=col_offset, min_row=row_offset + 1, max_row=data_end_row, max_col=col_offset)
        chart1.add_data(data, titles_from_data=False)
        chart1.set_categories(cats)
        
        # Add gap width and styling - simplified to avoid XML errors
        try:
            # Set chart colors using simpler approach
            from openpyxl.chart.series import DataPoint
            for series in chart1.series:
                # Use built-in chart style instead of custom colors to avoid XML issues
                pass  # Let Excel use default styling
        except Exception as e:
            st.warning(f"Chart styling warning: {e}")
        
        # Set gap width to match xlwings behavior (200% gap width)
        chart1.gapWidth = 200
        chart1.overlap = 0
        
        # Position chart with proper spacing - moved to column G
        chart1.width = 15
        chart1.height = 10
        ws.add_chart(chart1, f"G{chart_start_row}")
        
        # 2. User Ticket Completion Chart Data - Add more spacing between charts
        users_start_row = data_end_row + 15  # Increased spacing from 5 to 15
        ws.cell(row=users_start_row, column=col_offset, value="Ticket Completed by Individual")
        users_row_offset = users_start_row + 1
        ws.cell(row=users_row_offset, column=col_offset, value="Users")
        ws.cell(row=users_row_offset, column=col_offset + 1, value="Tickets")
        
        for i, (user, count) in enumerate(stats['ticket_completed'].items()):
            ws.cell(row=users_row_offset + i + 1, column=col_offset, value=user)
            ws.cell(row=users_row_offset + i + 1, column=col_offset + 1, value=count)
        
        # Create horizontal bar chart for users with styling
        chart2 = BarChart()
        chart2.type = "bar"
        chart2.style = 2  # Use same blue style for consistency
        chart2.title = "Users Completed"
        chart2.y_axis.title = 'Users'
        chart2.x_axis.title = 'Tickets'
        
        users_end_row = users_row_offset + len(stats['ticket_completed'])
        data2 = Reference(ws, min_col=col_offset + 1, min_row=users_row_offset + 1, max_row=users_end_row, max_col=col_offset + 1)
        cats2 = Reference(ws, min_col=col_offset, min_row=users_row_offset + 1, max_row=users_end_row, max_col=col_offset)
        chart2.add_data(data2, titles_from_data=False)
        chart2.set_categories(cats2)
        
        # Add gap width and styling - simplified to avoid XML errors
        try:
            # Use built-in chart styles for better compatibility
            for series in chart2.series:
                pass  # Let Excel use default styling
        except Exception as e:
            st.warning(f"Chart styling warning: {e}")
        
        chart2.gapWidth = 200
        chart2.overlap = 0
        chart2.width = 15
        chart2.height = 10
        ws.add_chart(chart2, f"G{users_start_row}")
        
        # 3. Priority Distribution Pie Chart - Add more spacing
        priority_start_row = users_end_row + 15  # Increased spacing from 5 to 15
        ws.cell(row=priority_start_row, column=col_offset, value="Priority wise ticket count")
        priority_row_offset = priority_start_row + 1
        ws.cell(row=priority_row_offset, column=col_offset, value="Priority")
        ws.cell(row=priority_row_offset, column=col_offset + 1, value="Count")
        
        for i, (priority, count) in enumerate(stats['priority'].items()):
            ws.cell(row=priority_row_offset + i + 1, column=col_offset, value=priority)
            ws.cell(row=priority_row_offset + i + 1, column=col_offset + 1, value=count)
        
        # Create pie chart for priority
        chart3 = PieChart()
        chart3.title = "Priority Distribution"
        priority_end_row = priority_row_offset + len(stats['priority'])
        data3 = Reference(ws, min_col=col_offset + 1, min_row=priority_row_offset + 1, max_row=priority_end_row, max_col=col_offset + 1)
        cats3 = Reference(ws, min_col=col_offset, min_row=priority_row_offset + 1, max_row=priority_end_row, max_col=col_offset)
        chart3.add_data(data3, titles_from_data=False)
        chart3.set_categories(cats3)
        chart3.width = 15
        chart3.height = 10
        ws.add_chart(chart3, f"G{priority_start_row}")
        
        # 4. SLA Chart Data - Add more spacing
        sla_start_row = priority_end_row + 15  # Increased spacing from 5 to 15
        ws.cell(row=sla_start_row, column=col_offset, value="SLA")
        sla_row_offset = sla_start_row + 1
        ws.cell(row=sla_row_offset, column=col_offset, value="SLA Status")
        ws.cell(row=sla_row_offset, column=col_offset + 1, value="Count")
        ws.cell(row=sla_row_offset + 1, column=col_offset, value="SLA MET")
        ws.cell(row=sla_row_offset + 1, column=col_offset + 1, value=100)
        ws.cell(row=sla_row_offset + 2, column=col_offset, value="SLA LOST")
        ws.cell(row=sla_row_offset + 2, column=col_offset + 1, value=0)
        
        # Create pie chart for SLA
        chart4 = PieChart()
        chart4.title = "SLA MET vs SLA LOST"
        data4 = Reference(ws, min_col=col_offset + 1, min_row=sla_row_offset + 1, max_row=sla_row_offset + 2, max_col=col_offset + 1)
        cats4 = Reference(ws, min_col=col_offset, min_row=sla_row_offset + 1, max_row=sla_row_offset + 2, max_col=col_offset)
        chart4.add_data(data4, titles_from_data=False)
        chart4.set_categories(cats4)
        chart4.width = 15
        chart4.height = 10
        ws.add_chart(chart4, f"G{sla_start_row}")
        
        # 5. Account Count Chart Data - Add more spacing
        account_start_row = sla_row_offset + 18  # Increased spacing from 8 to 18
        ws.cell(row=account_start_row, column=col_offset, value="Ticket Count by Accountwise")
        account_row_offset = account_start_row + 1
        ws.cell(row=account_row_offset, column=col_offset, value="Account")
        ws.cell(row=account_row_offset, column=col_offset + 1, value="Tickets")
        
        for i, (account, count) in enumerate(stats['account_count'].items()):
            ws.cell(row=account_row_offset + i + 1, column=col_offset, value=account)
            ws.cell(row=account_row_offset + i + 1, column=col_offset + 1, value=count)
        
        # Create horizontal bar chart for accounts with styling
        chart5 = BarChart()
        chart5.type = "bar"
        chart5.style = 2  # Use same blue style for consistency
        chart5.title = "Ticket Count by Accountwise"
        chart5.y_axis.title = 'Account'
        chart5.x_axis.title = 'Tickets'
        
        account_end_row = account_row_offset + len(stats['account_count'])
        data5 = Reference(ws, min_col=col_offset + 1, min_row=account_row_offset + 1, max_row=account_end_row, max_col=col_offset + 1)
        cats5 = Reference(ws, min_col=col_offset, min_row=account_row_offset + 1, max_row=account_end_row, max_col=col_offset)
        chart5.add_data(data5, titles_from_data=False)
        chart5.set_categories(cats5)
        
        # Add gap width and styling - simplified to avoid XML errors
        try:
            # Use built-in chart styles for better compatibility
            for series in chart5.series:
                pass  # Let Excel use default styling
        except Exception as e:
            st.warning(f"Chart styling warning: {e}")
        
        chart5.gapWidth = 200
        chart5.overlap = 0
        chart5.width = 15
        chart5.height = 10
        ws.add_chart(chart5, f"G{account_start_row}")
        
        # Save the file with preserved formatting and new charts
        output_path = st.session_state.file_path.replace('.xlsx', '_with_charts.xlsx')
        wb.save(output_path)
        
        # Create JSON data
        all_data = {
            'ticket_status_data': stats['dict_status'],
            'individual_data': stats['ticket_completed'], 
            'main_chart_data': stats['dict_status'],
            'pie1_data': stats['priority'],
            'pie2_data': {"SLA MET": 100, "SLA LOST": 0},
            'account_data': stats['account_count']
        }
        
        json_path = output_path.replace('.xlsx', '_data.json')
        with open(json_path, "w") as f:
            json.dump(all_data, f, indent=4)
        
        return output_path, json_path
        
    except Exception as e:
        st.error(f"Error generating charts with openpyxl: {e}")
        return None, None

def generate_json_data_only():
    """Generate JSON data without charts for fallback"""
    try:
        stats = st.session_state.stats
        
        # Create all_data structure similar to the xlwings version but without charts
        all_data = {
            'ticket_status_data': stats['dict_status'],
            'individual_data': stats['ticket_completed'], 
            'main_chart_data': stats['dict_status'],
            'pie1_data': stats['priority'],
            'pie2_data': {"SLA MET": 0, "SLA LOST": 0}  # Default SLA data
        }
        
        # Save JSON file
        json_path = st.session_state.file_path.replace('.xlsx', '_data.json')
        with open(json_path, "w") as f:
            json.dump(all_data, f, indent=4)
        
        return st.session_state.file_path, json_path
        
    except Exception as e:
        st.error(f"Error generating JSON data: {e}")
        return None, None

def generate_charts_and_save():
    """Generate charts with fallback: xlwings -> xlsxwriter -> JSON only"""
    try:
        # First try xlwings (works locally with Excel)
        try:
            st.info("Attempting to use xlwings for chart generation...")
            app = xw.App(visible=False)
            wb_xlwings = app.books.open(st.session_state.file_path)
            ws_xlwings = wb_xlwings.sheets['Cloud Services Report']
            chart_start_row = ws_xlwings.api.UsedRange.Rows.Count + 13
            stats = st.session_state.stats

            # Create data tables and charts - same approach as main.py but simplified
            # Ticket Status Chart Data
            ws_xlwings.range((chart_start_row, 5)).value = "TICKET STATUS"
            status_data = [["Status", "Count"]] + list(stats['dict_status'].items())
            ws_xlwings.range((chart_start_row + 1, 5)).value = status_data
            
            # User Ticket Completion Chart Data
            users_data_start_row = chart_start_row + len(status_data) +13
            ws_xlwings.range((users_data_start_row, 5)).value = "Ticket Completed by Individual"
            users_data = [["Users", "Tickets"]] + list(stats['ticket_completed'].items())
            ws_xlwings.range((users_data_start_row + 1, 5)).value = users_data

            # Priority Distribution Chart Data
            priority_data_start_row = users_data_start_row + len(users_data) + 13
            ws_xlwings.range((priority_data_start_row, 5)).value = "Priority wise ticket count"
            priority_data = [["Priority", "Count"]] + list(stats['priority'].items())
            ws_xlwings.range((priority_data_start_row + 1, 5)).value = priority_data

            # SLA Chart Data
            sla_data_start_row = priority_data_start_row + len(priority_data) + 13
            ws_xlwings.range((sla_data_start_row, 5)).value = "SLA"
            sla_data = [["SLA Met", "SLA Lost"]] + [[100, 0]]
            ws_xlwings.range((sla_data_start_row + 1, 5)).value = sla_data

            # Account Count Chart Data
            account_data_start_row = sla_data_start_row + len(sla_data) + 13
            ws_xlwings.range((account_data_start_row, 5)).value = "Ticket Count by Accountwise"
            account_data = [["Account", "Tickets"]] + list(stats['account_count'].items())
            ws_xlwings.range((account_data_start_row + 1, 5)).value = account_data
            
            # Create charts using the working chart functions
            add_horizontal_chart_xlwings(ws_xlwings, start_row=chart_start_row + 1, start_col=5, chart_title="Ticket Status Count", chart_index=0)
            add_horizontal_chart_xlwings(ws_xlwings, start_row=users_data_start_row + 1, start_col=5, chart_title="Users Completed", chart_index=1)
            add_pie_chart_xlwings(ws_xlwings, start_row=priority_data_start_row + 1, start_col=5, chart_title="Priority Distribution", chart_index=2)
            add_pie_chart_xlwings(ws_xlwings, start_row=sla_data_start_row + 1, start_col=5, chart_title="SLA MET vs SLA LOST", chart_index=3)
            add_horizontal_chart_xlwings(ws_xlwings, start_row=account_data_start_row + 1, start_col=5, chart_title="Ticket Count by Accountwise", chart_index=4)

            wb_xlwings.save()
            wb_xlwings.close()
            app.quit()

            # Export JSON data
            all_data = {
                "ticket_status": stats['dict_status'],
                "ticket_completed": stats['ticket_completed'],
                "priority_distribution": stats['priority'],
                "sla": stats['sla'],
                "account_count": stats['account_count']
            }
            
            json_path = st.session_state.file_path.replace('.xlsx', '_data.json')
            with open(json_path, "w") as f:
                json.dump(all_data, f, indent=4)
            
            st.success("Charts generated successfully using xlwings!")
            return st.session_state.file_path, json_path
            
        except Exception as xlwings_error:
            st.warning(f"xlwings failed: {xlwings_error}")
            st.info("Trying openpyxl with charts as fallback...")
            
            # Try openpyxl fallback (preserves styles and creates proper charts)
            try:
                excel_path, json_path = generate_charts_with_openpyxl()
                if excel_path and json_path:
                    st.success("Charts generated successfully using openpyxl!")
                    return excel_path, json_path
            except Exception as openpyxl_error:
                st.warning(f"openpyxl charts also failed: {openpyxl_error}")
            
            # Final fallback - JSON only
            st.info("Using JSON-only fallback...")
            return generate_json_data_only()
        
    except Exception as e:
        st.error(f"Error generating charts: {e}")
        return None, None

def create_combined_json_data():
    """Create combined JSON file with all processed data"""
    try:
        # Extract data from Excel processing
        stats = st.session_state.stats
        total_tasks = st.session_state.total
        completed_tasks = sum([v for k, v in stats['dict_status'].items() if k in ['Resolved with Customer', 'Closed']])
        
        # Get temp_daas data
        temp_daas_data = st.session_state.temp_daas_data
        
        # Prepare slide5 data from temp_daas processing
        slide5_data = None
        if temp_daas_data:
            slide5_data = {
                'summary_stats': {
                    'total_tickets': sum(temp_daas_data['status_counts'].values()),
                    'awaiting': temp_daas_data['status_counts'].get('Awaiting', 0),
                    'closed': temp_daas_data['status_counts'].get('Ticket closed', 0),
                    'resolved': temp_daas_data['status_counts'].get('Resolved with Customer', 0)
                },
                'daily_data': {}
            }
            
            # Convert date-wise data to daily format for slide5
            for date, data in temp_daas_data.get('date_wise_data', {}).items():
                slide5_data['daily_data'][date] = data.get('resources', {})
        
        # Get date information from extracted data
        date_info = st.session_state.get('date_info', {
            'period': '09/01/2025 to 09/07/2025',
            'report_date': '8 September 2025', 
            'new_date': '09/07/2025'
        })
        
        # Create comprehensive JSON structure
        # Main Excel data is used for slides 1-4 and slide 6
        # temp_daas_queue data is ONLY used for slide 5
        combined_data = {
            "metadata": {
                "report_date": date_info['report_date'],
                "new_period": date_info['period'],
                "new_date": date_info['new_date'],
                "generation_timestamp": int(time.time()),
                "total_tasks": total_tasks,
                "completed_tasks": completed_tasks,
                "extracted_from_b7": True
            },
            "main_report_data": {
                # Data for slides 1-4 comes from main Excel file
                "ticket_status_data": stats['dict_status'],  # Slide 2 chart 1
                "individual_data": stats['ticket_completed'], # Slide 2 chart 2  
                "main_chart_data": stats['account_count'],    # Slide 3 main chart
                "pie1_data": stats['sla'],                    # Slide 3 pie chart 1
                "pie2_data": stats['priority'],               # Slide 3 pie chart 2
                "all_stats": stats
            },
            # temp_daas_queue data - ONLY for slide 5
            "daas_queue_data": temp_daas_data,
            "slide5_data": slide5_data,  # This uses temp_daas_queue data
            "slide6_data": {
                # Slide 6 uses main Excel data for current week, historical data can be static
                'column_chart_data': {
                    "August 2nd Week": 180,
                    "August 3rd Week": 200, 
                    "August 4th Week": 220,
                    "September 1st Week": slide5_data["summary_stats"]["total_tickets"]  # From main Excel
                },
                'bar_chart_data': {
                    "August 2nd Week": {
                        'awaiting': 8,
                        'closed': 5,
                        'resolved': 187
                    },
                    "August 3rd Week": {
                        'awaiting': 12,
                        'closed': 3,
                        'resolved': 165
                    },
                    "August 4th Week": {
                        'awaiting': 12,
                        'closed': 3,
                        'resolved': 165
                    },
                    "September 1st Week": {
                        # Current week data from main Excel
                        'awaiting': slide5_data["summary_stats"]["awaiting"], 
                        'closed': slide5_data["summary_stats"]["closed"],
                        'resolved': slide5_data["summary_stats"]["resolved"]
                    }
                }
            }
        }
        
        # Save JSON file
        json_filename = f"combined_report_data_{int(time.time())}.json"
        with open(json_filename, 'w') as f:
            json.dump(combined_data, f, indent=4)
        
        # Store in session state for later use
        st.session_state.combined_json_data = combined_data
        st.session_state.combined_json_path = json_filename
        
        return json_filename, combined_data
        
    except Exception as e:
        st.error(f"Error creating combined JSON: {e}")
        return None, None

def generate_ppt_from_json(json_data):
    """Generate PowerPoint report using JSON data"""
    try:
        # Check if json_data is None
        if json_data is None:
            st.error("JSON data is None. Please generate the combined JSON data first.")
            return None
        
        # Check for required keys
        required_keys = ['metadata', 'main_report_data', 'slide5_data', 'slide6_data']
        missing_keys = [key for key in required_keys if key not in json_data]
        if missing_keys:
            st.error(f"Missing required keys in JSON data: {missing_keys}")
            st.write("Available keys:", list(json_data.keys()) if json_data else "None")
            return None
        
        # Check if template exists
        template_path = "template.pptx"
        if not os.path.exists(template_path):
            st.error(f"Template file '{template_path}' not found. Please ensure template.pptx is in the project directory.")
            return None
        
        output_path = f"final_report_{int(time.time())}.pptx"
        
        # Extract data from JSON
        metadata = json_data['metadata']
        main_data = json_data['main_report_data']
        
        # Generate PPT using JSON data
        generate_weekly_report(
            template_path=template_path,
            output_path=output_path,
            report_date=metadata['report_date'],
            new_period=metadata['new_period'],
            total_tasks=metadata['total_tasks'],
            completed_tasks=metadata['completed_tasks'],
            ticket_status_data=main_data['ticket_status_data'],
            individual_data=main_data['individual_data'],
            main_chart_data=main_data['main_chart_data'],
            pie1_data=main_data['pie1_data'],
            pie2_data=main_data['pie2_data'],
            new_date=metadata['new_date'],
            slide5_data=json_data['slide5_data'],
            slide6_data=json_data['slide6_data']
        )
        
        return output_path
        
    except Exception as e:
        st.error(f"Error generating PowerPoint from JSON: {e}")
        return None

# Main Streamlit App
def main():
    st.title("📊 CSM Report Processor")
    st.markdown("Process your Cloud Services Report with interactive charts generation")
    
    # Sidebar
    with st.sidebar:
        st.header("📋 Progress")
        if 'sections' in st.session_state:
            total_sections = len(st.session_state.sections) - 1
            current_section = st.session_state.current_section
            progress = (current_section / total_sections) if total_sections > 0 else 0
            st.progress(progress)
            st.write(f"Section {current_section + 1} of {total_sections}")
            st.write(f"Row {st.session_state.current_row_in_section} in current section")
            st.write(f"Deleted rows: {st.session_state.deleted_rows}")
        else:
            st.write("Upload a file to start processing")
        
        st.header("📊 Current Statistics")
        stats = st.session_state.stats
        
        # Show non-zero stats
        for category, data in stats.items():
            if any(v > 0 for v in data.values()):
                st.subheader(category.replace('_', ' ').title())
                for key, value in data.items():
                    if value > 0:
                        st.write(f"{key}: {value}")
    
    # Main content - Enhanced for dual file upload
    if not st.session_state.file_processed and not st.session_state.processing_complete:
        # File upload phase
        st.header("📁 Upload Excel Files")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.subheader("Main Report File")
            uploaded_file = st.file_uploader("Choose main Excel file (Cloud Services Report)", type=['xlsx', 'xls'], key="main_file")
        
        with col2:
            st.subheader("DaaS Queue File")
            temp_daas_file = st.file_uploader("Choose temp_daas_queue Excel file", type=['xlsx', 'xls'], key="daas_file")
        
        # Process files when both are uploaded
        if uploaded_file is not None:
            st.success("✅ Main report file uploaded")
            
            if temp_daas_file is not None:
                st.success("✅ DaaS queue file uploaded")
                
                if st.button("Start Processing Both Files", type="primary"):
                    with st.spinner("Processing uploaded files..."):
                        # Process main file
                        process_uploaded_file(uploaded_file)
                        
                        # Process temp_daas file in background
                        with st.spinner("Processing DaaS queue data..."):
                            temp_daas_data = process_temp_daas_file(temp_daas_file)
                            if temp_daas_data:
                                st.session_state.temp_daas_data = temp_daas_data
                                st.session_state.temp_daas_processed = True
                                st.success("DaaS queue data processed successfully!")
                            else:
                                st.warning("Using sample DaaS queue data")
                                st.session_state.temp_daas_data = {
                                    'resource_counts': create_sample_data()[0],
                                    'status_counts': create_sample_data()[1],
                                    'date_wise_data': create_sample_data()[2]
                                }
                                st.session_state.temp_daas_processed = True
                    
                    st.success("Files processed successfully!")
                    st.rerun()
            else:
                st.info("Please upload both files to continue")
        elif temp_daas_file is not None:
            st.info("Please upload the main report file as well")
    
    elif st.session_state.file_processed and not st.session_state.processing_complete:
        # Ticket processing phase using simplified logic from main.py
        st.header("Processing Tickets")
        
        current_ticket = get_current_ticket_for_processing()
        
        if current_ticket == "subtotal_found":
            st.session_state.r.append(st.session_state.current_row)
            st.session_state.current_row += 1
            st.info(f"Subtotal row found at row {st.session_state.current_row - 1}. Moving to the next ticket section.")
            st.rerun()
        elif current_ticket == "total_found":
            st.session_state.ws.cell(row=st.session_state.r[-1]+1, column=4, value=st.session_state.total)
            st.session_state.processing_complete = True
            st.success("All ticket rows processed!")
            st.rerun()
        elif current_ticket is not None:
            # Display current ticket 
            with st.form(key=f"form_{current_ticket['row']}"):
                st.subheader(f"Current Row: {current_ticket['row']}")
                st.write(f"**Ticket Status**: {status_str if current_ticket['status'] else current_ticket['status']}")
                st.write(f"**User**: {current_ticket['user']}")
                st.write(f"**Subject**: {current_ticket['subject']}")

                col1, col2 = st.columns(2)
                with col1:
                    action_text = st.text_input("Enter Action text:", placeholder="Required for update...")
                with col2:
                    account_options = list(st.session_state.stats['account_count'].keys())
                    selected_account = st.selectbox("Select Account:", account_options)
                
                # Side by side buttons
                button_col1, button_col2 = st.columns(2)
                with button_col1:
                    delete_row = st.form_submit_button("🗑️ Delete Row", use_container_width=True)
                with button_col2:
                    update_row = st.form_submit_button("✅ Update Row", use_container_width=True)

            if delete_row:
                process_current_ticket("delete")
                st.rerun()

            if update_row:
                # Validate action text before processing
                if not action_text or not action_text.strip():
                    st.error("❌ Action text is required for updating the row!")
                else:
                    process_current_ticket("update", action_text, selected_account)
                    st.rerun()
        else:
            # All tickets processed
            st.session_state.processing_complete = True
            st.rerun()
    
    elif st.session_state.processing_complete:
        # Results phase with enhanced report generation section
        st.header("🎉 Processing Complete!")
        st.success("All tickets have been processed successfully!")
        
        # Summary statistics at the top
        col1, col2, col3, col4 = st.columns(4)
        total_processed = st.session_state.get('total', 0)
        stats = st.session_state.stats
        
        with col1:
            st.markdown("""
            <div class="stat-card">
                <h3>📊 Total Tickets</h3>
                <h2>{}</h2>
            </div>
            """.format(total_processed), unsafe_allow_html=True)
        
        with col2:
            completed_count = sum([v for k, v in stats['dict_status'].items() if k in ['Resolved with Customer', 'Closed']])
            st.markdown("""
            <div class="stat-card">
                <h3>✅ Completed</h3>
                <h2>{}</h2>
            </div>
            """.format(completed_count), unsafe_allow_html=True)
        
        with col3:
            pending_count = sum([v for k, v in stats['dict_status'].items() if k in ['New', 'Inprogress', 'Awaiting']])
            st.markdown("""
            <div class="stat-card">
                <h3>⏳ Pending</h3>
                <h2>{}</h2>
            </div>
            """.format(pending_count), unsafe_allow_html=True)
        
        with col4:
            completion_rate = (completed_count / total_processed * 100) if total_processed > 0 else 0
            st.markdown("""
            <div class="stat-card">
                <h3>📈 Completion Rate</h3>
                <h2>{:.0f}%</h2>
            </div>
            """.format(completion_rate), unsafe_allow_html=True)
        
        st.markdown("---")
        
        # Enhanced Report generation section
        st.markdown('<div class="report-card">', unsafe_allow_html=True)
        st.markdown("### 📋 Report Generation")
        st.markdown("Generate various report formats from your processed data")
        
        # Three main action cards
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.markdown("#### 📊 Excel Report")
            st.markdown("Generate Excel file with charts and visualizations")
            
            if st.button("🚀 Generate Charts & Excel", type="primary", use_container_width=True):
                # Create modal dialog using popover
                with st.expander("📊 Generating Excel Report...", expanded=True):
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    
                    status_text.text("🔄 Starting chart generation...")
                    progress_bar.progress(10)
                    time.sleep(0.5)
                    
                    status_text.text("📊 Creating charts...")
                    progress_bar.progress(30)
                    excel_path, json_path = generate_charts_and_save()
                    
                    progress_bar.progress(80)
                    status_text.text("💾 Saving files...")
                    time.sleep(0.5)
                    
                    progress_bar.progress(100)
                    status_text.text("✅ Excel report generated successfully!")
                
                # Store in session state for persistence
                if excel_path and json_path:
                    st.session_state.excel_generated = True
                    st.session_state.excel_path = excel_path
                    st.session_state.json_path = json_path
                    st.success("📊 Charts generated successfully!")
                    st.rerun()
            
            # Show download buttons if Excel is generated
            if st.session_state.get('excel_generated', False) and st.session_state.get('excel_path'):
                try:
                    with open(st.session_state.excel_path, "rb") as file:
                        st.download_button(
                            label="📥 Download Excel Report",
                            data=file.read(),
                            file_name=f"processed_report_{int(time.time())}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                    
                    with open(st.session_state.json_path, "rb") as file:
                        st.download_button(
                            label="📄 Download JSON Data",
                            data=file.read(),
                            file_name=f"report_data_{int(time.time())}.json",
                            mime="application/json",
                            use_container_width=True
                        )
                except FileNotFoundError:
                    st.warning("Files not found. Please regenerate.")
        
        with col2:
            st.markdown("#### 🔗 Combined JSON")
            st.markdown("Create comprehensive JSON with all processed data")
            
            if st.button("🔗 Generate Combined JSON", type="primary", use_container_width=True):
                if st.session_state.temp_daas_processed:
                    with st.expander("🔗 Generating Combined JSON...", expanded=True):
                        progress_bar = st.progress(0)
                        status_text = st.empty()
                        
                        status_text.text("🔄 Starting JSON generation...")
                        progress_bar.progress(20)
                        time.sleep(0.3)
                        
                        status_text.text("🔗 Combining data sources...")
                        progress_bar.progress(50)
                        json_path, json_data = create_combined_json_data()
                        
                        progress_bar.progress(80)
                        status_text.text("💾 Saving combined JSON...")
                        time.sleep(0.3)
                        
                        progress_bar.progress(100)
                        status_text.text("✅ Combined JSON created!")
                    
                    # Store in session state for persistence and PPT generation
                    if json_path and json_data:
                        st.session_state.combined_json_generated = True
                        st.session_state.combined_json_path = json_path
                        st.session_state.combined_json_data = json_data
                        st.success("🔗 Combined JSON created successfully!")
                        st.rerun()
                else:
                    st.error("❌ Please process DaaS queue file first")
            
            # Show download button if Combined JSON is generated
            if st.session_state.get('combined_json_generated', False) and st.session_state.get('combined_json_path'):
                try:
                    with open(st.session_state.combined_json_path, "rb") as file:
                        st.download_button(
                            label="📥 Download Combined JSON",
                            data=file.read(),
                            file_name=f"combined_data_{int(time.time())}.json",
                            mime="application/json",
                            use_container_width=True
                        )
                    
                    # Show JSON preview in an expandable section
                    with st.expander("🔍 Preview JSON Structure"):
                        st.json(st.session_state.combined_json_data['metadata'])
                except FileNotFoundError:
                    st.warning("JSON file not found. Please regenerate.")
        
        with col3:
            st.markdown("#### 🎯 PowerPoint Report")
            st.markdown("Generate final presentation from combined data")
            
            if st.button("🎯 Generate PowerPoint", type="primary", use_container_width=True):
                # Auto-generate JSON if not already generated
                if not st.session_state.get('combined_json_generated', False):
                    if st.session_state.temp_daas_processed:
                        with st.expander("🔗 Auto-generating Combined JSON...", expanded=True):
                            progress_bar = st.progress(0)
                            status_text = st.empty()
                            
                            status_text.text("🔄 Auto-generating JSON for PPT...")
                            progress_bar.progress(25)
                            json_path, json_data = create_combined_json_data()
                            
                            if json_path and json_data:
                                st.session_state.combined_json_generated = True
                                st.session_state.combined_json_path = json_path
                                st.session_state.combined_json_data = json_data
                                progress_bar.progress(50)
                                status_text.text("✅ JSON auto-generated successfully!")
                    else:
                        st.error("❌ Please process DaaS queue file first")
                        st.stop()
                
                # Generate PowerPoint
                if st.session_state.get('combined_json_data'):
                    with st.expander("🎯 Generating PowerPoint...", expanded=True):
                        progress_bar = st.progress(0)
                        status_text = st.empty()
                        
                        status_text.text("🎯 Creating PowerPoint slides...")
                        progress_bar.progress(30)
                        time.sleep(0.5)
                        
                        ppt_path = generate_ppt_from_json(st.session_state.combined_json_data)
                        
                        progress_bar.progress(80)
                        status_text.text("💾 Saving PowerPoint...")
                        time.sleep(0.5)
                        
                        progress_bar.progress(100)
                        status_text.text("✅ PowerPoint generated!")
                    
                    # Store in session state for persistence
                    if ppt_path and os.path.exists(ppt_path):
                        st.session_state.ppt_generated = True
                        st.session_state.ppt_path = ppt_path
                        st.success("🎯 PowerPoint report generated successfully!")
                        st.rerun()
            
            # Show download button if PPT is generated
            if st.session_state.get('ppt_generated', False) and st.session_state.get('ppt_path'):
                try:
                    with open(st.session_state.ppt_path, "rb") as file:
                        st.download_button(
                            label="📥 Download PowerPoint",
                            data=file.read(),
                            file_name=f"final_report_{int(time.time())}.pptx",
                            mime="application/vnd.openxmlformats-officedocument.presentationml.presentation",
                            use_container_width=True
                        )
                except FileNotFoundError:
                    st.warning("PowerPoint file not found. Please regenerate.")
        
        st.markdown('</div>', unsafe_allow_html=True)
        
        # Reset option
        st.markdown("---")
        col1, col2, col3 = st.columns([1, 1, 1])
        with col2:
            if st.button("🔄 Process New Files", type="secondary", use_container_width=True):
                # Cleanup temp files before reset
                cleanup_temp_files()
                
                # Cleanup generated files
                if st.session_state.get('excel_path') and os.path.exists(st.session_state.excel_path):
                    try:
                        os.remove(st.session_state.excel_path)
                    except:
                        pass
                        
                if st.session_state.get('combined_json_path') and os.path.exists(st.session_state.combined_json_path):
                    try:
                        os.remove(st.session_state.combined_json_path)
                    except:
                        pass
                        
                if st.session_state.get('ppt_path') and os.path.exists(st.session_state.ppt_path):
                    try:
                        os.remove(st.session_state.ppt_path)
                    except:
                        pass
                
                # Reset session state
                for key in list(st.session_state.keys()):
                    del st.session_state[key]
                st.rerun()
        
        # Display final statistics
        st.header("📈 Final Statistics")
        
        stats = st.session_state.stats
        
        # Main report statistics
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.subheader("🎯 Status Distribution")
            status_data = {k: v for k, v in stats['dict_status'].items() if v > 0}
            if status_data:
                st.bar_chart(status_data)
        
        with col2:
            st.subheader("⭐ Priority Distribution") 
            priority_data = {k: v for k, v in stats['priority'].items() if v > 0}
            if priority_data:
                st.bar_chart(priority_data)
        
        with col3:
            st.subheader("👥 User Performance")
            user_data = {k: v for k, v in stats['ticket_completed'].items() if v > 0}
            if user_data:
                st.bar_chart(user_data)
        
        # DaaS Queue statistics if available
        if st.session_state.temp_daas_processed and st.session_state.temp_daas_data:
            st.header("📊 DaaS Queue Analysis")
            
            temp_daas_data = st.session_state.temp_daas_data
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.subheader("🔧 Resource Distribution")
                resource_data = {k: v for k, v in temp_daas_data['resource_counts'].items() if v > 0}
                if resource_data:
                    st.bar_chart(resource_data)
            
            with col2:
                st.subheader("📈 Status Breakdown")
                daas_status_data = {k: v for k, v in temp_daas_data['status_counts'].items() if v > 0}
                if daas_status_data:
                    st.bar_chart(daas_status_data)
            
            # Date-wise analysis
            if temp_daas_data.get('date_wise_data'):
                st.subheader("📅 Date-wise Analysis")
                date_wise_df = []
                for date, data in temp_daas_data['date_wise_data'].items():
                    total_resources = sum(data.get('resources', {}).values())
                    date_wise_df.append({'Date': date, 'Total Tickets': total_resources})
                
                if date_wise_df:
                    import pandas as pd
                    df = pd.DataFrame(date_wise_df)
                    st.line_chart(df.set_index('Date'))
    

if __name__ == "__main__":
    # Cleanup temp files on app exit using atexit
    import atexit
    
    def cleanup_on_exit():
        """Cleanup function called when app exits"""
        if 'temp_files' in st.session_state:
            for temp_file in st.session_state.temp_files:
                try:
                    if os.path.exists(temp_file):
                        os.remove(temp_file)
                except:
                    pass  # Silent cleanup on exit
    
    atexit.register(cleanup_on_exit)
    
    # Add manual cleanup button in sidebar for immediate cleanup
    with st.sidebar:
        if st.button("🧹 Cleanup Temp Files", help="Manually remove temporary files"):
            cleanup_temp_files()
    
    main()